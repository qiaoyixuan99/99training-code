"""
Spare Parts ABC Classification Module (备件ABC等级分类模块)
==========================================================

Grade assignment based on composite Sum score:
  Sum = Q*0.1 + S*0.1 + V*0.2 + D*0.3 + P*0.3

Where:
  Q — Quality (质量)        weight: 0.1
  S — Safety (安全)         weight: 0.1
  V — Value (价格)          weight: 0.2
  D — Delivery (交期)       weight: 0.3
  P — Equipment importance (所在设备重要性) weight: 0.3

Classification thresholds (derived from over3 data analysis):
  Sum >= 4.70  →  Grade A (高优先级)
  4.66 <= Sum < 4.70  →  Grade B (中优先级)
  Sum < 4.66  →  Grade C (低优先级)

Usage:
  from classify_parts import classify_by_sum, classify_by_s, compute_sum

  grade = classify_by_sum(q=5, s=4, v=3, d=5, p=5)
  grade = classify_by_s(s=4.5)
  total = compute_sum(q=5, s=4, v=3, d=5, p=5)
"""

# Sum weight coefficients
W_Q = 0.1  # Quality
W_S = 0.1  # Safety
W_V = 0.2  # Value
W_D = 0.3  # Delivery
W_P = 0.3  # Equipment importance

# Sum-based classification thresholds
# Derived from S-based grading applied to over3 data (3593 rows)
# Grade A (S>4.5): mean Sum=4.72
# Grade B (2.5<S<4.5): mean Sum=4.68
# Grade C (S<2.5): mean Sum=4.63
THRESHOLD_AB = 4.70   # A/B boundary: (4.72 + 4.68) / 2
THRESHOLD_BC = 4.66   # B/C boundary: (4.68 + 4.63) / 2


def compute_sum(q, s, v, d, p):
    """
    Compute weighted Sum score from the 5 evaluation factors.

    Args:
        q: Quality score (质量, 1-5)
        s: Safety score (安全, 1-5)
        v: Value score (价格, 1-5)
        d: Delivery score (交期, 1-5)
        p: Equipment importance score (所在设备重要性, 1-5)

    Returns:
        float: Composite sum = Q*0.1 + S*0.1 + V*0.2 + D*0.3 + P*0.3
    """
    return q * W_Q + s * W_S + v * W_V + d * W_D + p * W_P


def classify_by_sum(q, s, v, d, p):
    """
    Classify spare part into ABC grade based on composite Sum score.

    This is the PRIMARY function for new data entries.
    Computes Sum from the 5 factors and returns the corresponding grade.

    Args:
        q: Quality score (1-5)
        s: Safety score (1-5)
        v: Value score (1-5)
        d: Delivery score (1-5)
        p: Equipment importance score (1-5)

    Returns:
        str: 'A', 'B', or 'C'

    Example:
        >>> classify_by_sum(q=5, s=5, v=5, d=5, p=5)
        'A'
        >>> classify_by_sum(q=3, s=3, v=3, d=3, p=3)
        'C'
    """
    total = compute_sum(q, s, v, d, p)
    if total >= THRESHOLD_AB:
        return 'A'
    elif total >= THRESHOLD_BC:
        return 'B'
    else:
        return 'C'


def classify_by_s(s):
    """
    Classify spare part into ABC grade based on Safety (S) value alone.

    Rules:
        S > 4.5  →  A
        2.5 < S < 4.5  →  B
        S < 2.5  →  C

    Args:
        s: Safety score (安全)

    Returns:
        str: 'A', 'B', or 'C'
    """
    if s is None:
        return 'C'
    s = float(s)
    if s > 4.5:
        return 'A'
    elif s > 2.5:
        return 'B'
    else:
        return 'C'


def classify_both(q, s, v, d, p):
    """
    Classify using both methods and return detailed results.

    Args:
        q, s, v, d, p: The 5 evaluation factor scores (1-5)

    Returns:
        dict: {
            'sum': float,
            'grade_by_sum': str,
            'grade_by_s': str,
            's_value': float/int,
        }
    """
    total = compute_sum(q, s, v, d, p)
    return {
        'sum': round(total, 2),
        'grade_by_sum': classify_by_sum(q, s, v, d, p),
        'grade_by_s': classify_by_s(s),
        's_value': s,
    }


# ============================================================
# Batch processing helper
# ============================================================
def process_excel(input_path, output_path):
    """
    Read an Excel file, compute Sum and ABC grade for each row,
    and save to a new file.

    Expects columns: AA(Q), AB(S), AC(V), AD(D), AE(P)
    Writes to: AF(Sum formula), L(Class)

    Args:
        input_path: Path to input .xlsx file
        output_path: Path to output .xlsx file
    """
    import openpyxl

    COL_CLASS = 12   # L
    COL_Q     = 27   # AA
    COL_S     = 28   # AB
    COL_V     = 29   # AC
    COL_D     = 30   # AD
    COL_P     = 31   # AE
    COL_SUM   = 32   # AF
    COL_MODEL = 8    # H

    wb = openpyxl.load_workbook(input_path)
    ws = wb['Inventory']

    count = 0
    empty_streak = 0
    for row_idx in range(2, ws.max_row + 1):
        model = ws.cell(row=row_idx, column=COL_MODEL).value
        if model is None:
            empty_streak += 1
            if empty_streak >= 200:
                break
            continue
        empty_streak = 0

        q = ws.cell(row=row_idx, column=COL_Q).value or 0
        s = ws.cell(row=row_idx, column=COL_S).value or 0
        v = ws.cell(row=row_idx, column=COL_V).value or 0
        d = ws.cell(row=row_idx, column=COL_D).value or 0
        p = ws.cell(row=row_idx, column=COL_P).value or 0

        # Write Sum as Excel formula
        formula = f"=AA{row_idx}*0.1+AB{row_idx}*0.1+AC{row_idx}*0.2+AD{row_idx}*0.3+AE{row_idx}*0.3"
        ws.cell(row=row_idx, column=COL_SUM).value = formula

        # Write grade based on Sum
        grade = classify_by_sum(q, s, v, d, p)
        ws.cell(row=row_idx, column=COL_CLASS).value = grade
        count += 1

    wb.save(output_path)
    print(f"Processed {count} rows. Saved to: {output_path}")
    return count


# ============================================================
# Command-line entry point
# ============================================================
if __name__ == '__main__':
    import sys

    if len(sys.argv) >= 3:
        process_excel(sys.argv[1], sys.argv[2])
    else:
        # Quick demo
        print("ABC Classification Demo")
        print("=" * 50)
        test_cases = [
            (5, 5, 5, 5, 5),
            (4, 4, 4, 4, 4),
            (3, 3, 3, 3, 3),
            (2, 2, 2, 2, 2),
            (1, 1, 1, 1, 1),
            (5, 3, 5, 5, 5),
            (3, 5, 3, 3, 3),
        ]
        print(f"{'Q':>3} {'S':>3} {'V':>3} {'D':>3} {'P':>3}  {'Sum':>6}  {'Sum-Grade':>10}  {'S-Grade':>8}")
        print("-" * 60)
        for q, s, v, d, p in test_cases:
            result = classify_both(q, s, v, d, p)
            print(f"{q:3} {s:3} {v:3} {d:3} {p:3}  {result['sum']:6.2f}  {result['grade_by_sum']:>10}  {result['grade_by_s']:>8}")
        print()
        print(f"Thresholds: A/B={THRESHOLD_AB}, B/C={THRESHOLD_BC}")
        print(f"Formula: Sum = Q*{W_Q} + S*{W_S} + V*{W_V} + D*{W_D} + P*{W_P}")
        print()
        print("Usage: python classify_parts.py <input.xlsx> <output.xlsx>")
