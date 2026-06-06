"""
Process change1.xlsx -> over2.xlsx

Operations:
1. Slim down IN/OUT sheets: delete rows with date < 2026-01-01
2. Adjust Inventory Original quantity (O) to compensate for deleted IN/OUT data,
   so current stock quantities (P) remain unchanged.
3. Inventory sheet: set uniqueness(Z)=0, fill Q/S/V/D/P(AA-AE) with random 1-5
   integers ensuring weighted sum > 4.5, write Excel formula in AF(Sum),
   set class(L)="A", fill class cell red if stock quantity < 2.
4. Preserve table structure — no column/row resizing, no format changes beyond specified.
"""
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime
from collections import defaultdict
import random
import copy

SRC = r'd:\【99】training code\make money\1-库存管理\change1.xlsx'
DST = r'd:\【99】training code\make money\1-库存管理\over2.xlsx'

CUTOFF = datetime(2026, 1, 1)

# ============================================================
# Phase 1: Read current data (data_only mode) to capture
#          computed inventory quantities before any changes
# ============================================================
print("[1/5] Reading current data...")
wb_data = openpyxl.load_workbook(SRC, data_only=True)

ws_inv_data = wb_data['Inventory']
ws_in_data = wb_data['IN']
ws_out_data = wb_data['OUT']

# Build lookup: for each inventory row, get current computed P (Quantity),
# H (Model), and O (Original quantity)
# Inventory columns (1-indexed): H=8 (Model), O=15 (Original qty), P=16 (Qty)
# NOTE: ws_inv_data.max_row may be 1048575 — we stop after 200 consecutive empty rows
# to avoid iterating the entire sheet.
inv_rows_info = []  # list of dicts: {row, model, original, current_qty}
empty_streak = 0
for row_idx in range(2, ws_inv_data.max_row + 1):
    model = ws_inv_data.cell(row=row_idx, column=8).value   # H
    qty = ws_inv_data.cell(row=row_idx, column=16).value     # P (computed)
    original = ws_inv_data.cell(row=row_idx, column=15).value # O

    if model is not None or qty is not None:
        inv_rows_info.append({
            'row': row_idx,
            'model': str(model).strip() if model else '',
            'original': float(original) if original else 0.0,
            'current_qty': float(qty) if qty else 0.0,
        })
        empty_streak = 0
    else:
        empty_streak += 1
        if empty_streak >= 200:
            break  # reached the end of actual data

print(f"  Inventory data rows: {len(inv_rows_info)}")

# Build IN sums (2026+) per model
# IN columns: F=6 (Model), M=13 (Qty In), O=15 (Date)
in_2026 = defaultdict(float)
in_total = defaultdict(float)
for row_idx in range(2, ws_in_data.max_row + 1):
    model = ws_in_data.cell(row=row_idx, column=6).value   # F
    qty = ws_in_data.cell(row=row_idx, column=13).value     # M
    date = ws_in_data.cell(row=row_idx, column=15).value    # O
    if model is None:
        continue
    model_key = str(model).strip()
    qty_val = float(qty) if qty else 0.0
    in_total[model_key] += qty_val
    if isinstance(date, datetime) and date >= CUTOFF:
        in_2026[model_key] += qty_val

# Build OUT sums (2026+) per model
# OUT columns: F=6 (Model), M=13 (Qty Out), N=14 (Date)
out_2026 = defaultdict(float)
out_total = defaultdict(float)
for row_idx in range(2, ws_out_data.max_row + 1):
    model = ws_out_data.cell(row=row_idx, column=6).value   # F
    qty = ws_out_data.cell(row=row_idx, column=13).value     # M
    date = ws_out_data.cell(row=row_idx, column=14).value    # N
    if model is None:
        continue
    model_key = str(model).strip()
    qty_val = float(qty) if qty else 0.0
    out_total[model_key] += qty_val
    if isinstance(date, datetime) and date >= CUTOFF:
        out_2026[model_key] += qty_val

print(f"  IN models: {len(in_total)}, OUT models: {len(out_total)}")
print(f"  IN 2026+ models: {len(in_2026)}, OUT 2026+ models: {len(out_2026)}")

wb_data.close()

# ============================================================
# Phase 2: Load editable workbook
# ============================================================
print("\n[2/5] Loading editable workbook...")
wb = openpyxl.load_workbook(SRC)
ws_inv = wb['Inventory']
ws_in = wb['IN']
ws_out = wb['OUT']

# ============================================================
# Phase 3: Adjust Inventory Original quantity (O) to preserve
#          current stock after IN/OUT old data deletion.
#          Formula: P = O + SUMIF(IN) - SUMIF(OUT)
#          After deletion, SUMIF only covers 2026+ data.
#          New_O = Current_P - IN_2026 + OUT_2026
# ============================================================
print("\n[3/5] Adjusting Inventory Original quantities...")
adjusted_count = 0
for info in inv_rows_info:
    row_idx = info['row']
    model = info['model']
    current_qty = info['current_qty']
    in26 = in_2026.get(model, 0.0)
    out26 = out_2026.get(model, 0.0)
    new_original = current_qty - in26 + out26

    o_cell = ws_inv.cell(row=row_idx, column=15)  # O column
    o_cell.value = new_original
    if new_original != info['original']:
        adjusted_count += 1

print(f"  Rows with adjusted Original quantity: {adjusted_count}")

# ============================================================
# Phase 4: Delete IN/OUT rows with date < 2026-01-01
#          Use batch deletion: group consecutive rows and delete each
#          block in one call, avoiding O(n²) from row-by-row deletion.
# ============================================================
def batch_delete_rows(ws, col, cutoff):
    """Delete rows where column `col` has a datetime < cutoff.
    Groups consecutive rows and deletes each block in one operation."""
    to_delete = []
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=col).value
        if isinstance(val, datetime) and val < cutoff:
            to_delete.append(row_idx)

    if not to_delete:
        return 0

    # Group consecutive row numbers into (start, end) ranges
    ranges = []
    start = to_delete[0]
    end = to_delete[0]
    for r in to_delete[1:]:
        if r == end + 1:
            end = r
        else:
            ranges.append((start, end))
            start = end = r
    ranges.append((start, end))

    # Delete from bottom to top (preserves upper row indices)
    for start, end in reversed(ranges):
        ws.delete_rows(start, end - start + 1)

    return len(to_delete)

print("\n[4/5] Deleting IN/OUT rows before 2026...")

# IN sheet — date in column O (col 15)
in_deleted = batch_delete_rows(ws_in, col=15, cutoff=CUTOFF)
print(f"  IN rows deleted: {in_deleted}, remaining: {ws_in.max_row}")

# OUT sheet — date in column N (col 14)
out_deleted = batch_delete_rows(ws_out, col=14, cutoff=CUTOFF)
print(f"  OUT rows deleted: {out_deleted}, remaining: {ws_out.max_row}")

# ============================================================
# Phase 5: Update Inventory columns Z-AF and Class
# ============================================================
print("\n[5/5] Updating Inventory scores and class...")

COL_UNIQUE = 26   # Z  — 唯一性
COL_Q      = 27   # AA — 质量 Quality
COL_S      = 28   # AB — 安全 Safety
COL_V      = 29   # AC — 价格 Value
COL_D      = 30   # AD — 交期 Delivery
COL_P_EQ   = 31   # AE — 所在设备重要性 Equipment importance
COL_SUM    = 32   # AF — 分值 Sum (formula)
COL_CLASS  = 12   # L  — 等级 class
COL_QTY    = 16   # P  — 库存数量 Quantity

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

rows_scored = 0
rows_red = 0

for info in inv_rows_info:
    row_idx = info['row']
    qty = info['current_qty']  # Use the current (preserved) quantity

    # Generate random 1-5 integers; reject until weighted sum > 4.5
    # Weights: Q*0.1 + S*0.1 + V*0.2 + D*0.3 + P*0.3
    # D and P contribute most (0.3 each), V next (0.2), Q/S least (0.1)
    for _ in range(10000):
        q_val = random.randint(1, 5)
        s_val = random.randint(1, 5)
        v_val = random.randint(1, 5)
        d_val = random.randint(1, 5)
        p_val = random.randint(1, 5)
        if q_val*0.1 + s_val*0.1 + v_val*0.2 + d_val*0.3 + p_val*0.3 > 4.5:
            break

    # Write integer values
    ws_inv.cell(row=row_idx, column=COL_UNIQUE).value = 0
    ws_inv.cell(row=row_idx, column=COL_Q).value = q_val
    ws_inv.cell(row=row_idx, column=COL_S).value = s_val
    ws_inv.cell(row=row_idx, column=COL_V).value = v_val
    ws_inv.cell(row=row_idx, column=COL_D).value = d_val
    ws_inv.cell(row=row_idx, column=COL_P_EQ).value = p_val

    # Excel formula for Sum: =AA*0.1+AB*0.1+AC*0.2+AD*0.3+AE*0.3
    formula = f"=AA{row_idx}*0.1+AB{row_idx}*0.1+AC{row_idx}*0.2+AD{row_idx}*0.3+AE{row_idx}*0.3"
    ws_inv.cell(row=row_idx, column=COL_SUM).value = formula

    # Set Class to "A"
    class_cell = ws_inv.cell(row=row_idx, column=COL_CLASS)
    class_cell.value = "A"

    # Red fill if stock quantity < 2
    if qty < 2:
        class_cell.fill = red_fill
        rows_red += 1

    rows_scored += 1

print(f"  Rows scored: {rows_scored}")
print(f"  Rows with red class (qty < 2): {rows_red}")

# ============================================================
# Save
# ============================================================
print(f"\nSaving to: {DST}")
wb.save(DST)
print("Done! over2.xlsx created successfully.")
