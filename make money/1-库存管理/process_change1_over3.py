"""
Process change1.xlsx -> over3.xlsx

Operations (extends over2 logic):
1. Slim down IN/OUT sheets: delete rows with date < 2026-01-01 (batch delete)
2. Adjust Inventory Original quantity (O) to compensate for deleted IN/OUT data,
   so current stock quantities (P) remain unchanged.
3. Uncollapse/ungroup all row outlines in IN and OUT sheets (取消区域折叠)
4. Delete truly empty rows in IN and OUT sheets (清除无数据的空白行)
5. Inventory sheet: set uniqueness(Z)=0, fill Q/S/V/D/P(AA-AE) with random 1-5
   integers ensuring weighted sum > 4.5, write Excel formula in AF(Sum),
   set class(L)="A", fill class cell red if stock quantity < 2.
6. Preserve table structure — no column/row resizing, no format changes beyond specified.
"""
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.worksheet.dimensions import RowDimension
from datetime import datetime
from collections import defaultdict
import random

SRC = r'd:\【99】training code\make money\1-库存管理\change1.xlsx'
DST = r'd:\【99】training code\make money\1-库存管理\over3.xlsx'

CUTOFF = datetime(2026, 1, 1)

# ============================================================
# Phase 1: Read current data (data_only mode) to capture
#          computed inventory quantities before any changes
# ============================================================
print("[1/7] Reading current data...")
wb_data = openpyxl.load_workbook(SRC, data_only=True)

ws_inv_data = wb_data['Inventory']
ws_in_data = wb_data['IN']
ws_out_data = wb_data['OUT']

# Build lookup: for each inventory row, get current computed P (Quantity),
# H (Model), and O (Original quantity)
# NOTE: ws_inv_data.max_row may be 1048575 — we stop after 200 consecutive empty rows
inv_rows_info = []
empty_streak = 0
for row_idx in range(2, ws_inv_data.max_row + 1):
    model = ws_inv_data.cell(row=row_idx, column=8).value     # H — Model
    qty = ws_inv_data.cell(row=row_idx, column=16).value       # P — Quantity (computed)
    original = ws_inv_data.cell(row=row_idx, column=15).value  # O — Original

    if model is not None or qty is not None:
        inv_rows_info.append({
            'row': row_idx,
            'model': str(model).strip() if model else '',
            'original': float(original) if original else 0.0,
            'current_qty': float(qty) if qty else 0.0,
        })
        empty_streak = 0
    else:
        empty_streak += 1
        if empty_streak >= 200:
            break

print(f"  Inventory data rows: {len(inv_rows_info)}")

# Build IN sums (2026+) per model
# IN columns: F=6 (Model), M=13 (Qty In), O=15 (Date)
in_2026 = defaultdict(float)
for row_idx in range(2, ws_in_data.max_row + 1):
    model = ws_in_data.cell(row=row_idx, column=6).value
    qty = ws_in_data.cell(row=row_idx, column=13).value
    date = ws_in_data.cell(row=row_idx, column=15).value
    if model is None:
        continue
    model_key = str(model).strip()
    qty_val = float(qty) if qty else 0.0
    if isinstance(date, datetime) and date >= CUTOFF:
        in_2026[model_key] += qty_val

# Build OUT sums (2026+) per model
# OUT columns: F=6 (Model), M=13 (Qty Out), N=14 (Date)
out_2026 = defaultdict(float)
for row_idx in range(2, ws_out_data.max_row + 1):
    model = ws_out_data.cell(row=row_idx, column=6).value
    qty = ws_out_data.cell(row=row_idx, column=13).value
    date = ws_out_data.cell(row=row_idx, column=14).value
    if model is None:
        continue
    model_key = str(model).strip()
    qty_val = float(qty) if qty else 0.0
    if isinstance(date, datetime) and date >= CUTOFF:
        out_2026[model_key] += qty_val

print(f"  IN 2026+ models: {len(in_2026)}, OUT 2026+ models: {len(out_2026)}")
wb_data.close()

# ============================================================
# Phase 2: Load editable workbook
# ============================================================
print("\n[2/7] Loading editable workbook...")
wb = openpyxl.load_workbook(SRC)
ws_inv = wb['Inventory']
ws_in = wb['IN']
ws_out = wb['OUT']

# ============================================================
# Phase 3: Adjust Inventory Original quantity (O) to preserve
#          current stock after IN/OUT old data deletion.
#          Formula: P = O + SUMIF(IN) - SUMIF(OUT)
#          After deletion, SUMIF only covers 2026+ data.
#          New_O = Current_P - IN_2026 + OUT_2026
# ============================================================
print("\n[3/7] Adjusting Inventory Original quantities...")
adjusted_count = 0
for info in inv_rows_info:
    row_idx = info['row']
    model = info['model']
    current_qty = info['current_qty']
    in26 = in_2026.get(model, 0.0)
    out26 = out_2026.get(model, 0.0)
    new_original = current_qty - in26 + out26

    o_cell = ws_inv.cell(row=row_idx, column=15)
    o_cell.value = new_original
    if new_original != info['original']:
        adjusted_count += 1

print(f"  Rows with adjusted Original quantity: {adjusted_count}")

# ============================================================
# Phase 4: Delete IN/OUT rows with date < 2026-01-01 (batch)
# ============================================================
def batch_delete_rows(ws, col, cutoff):
    """Delete rows where column `col` has a datetime < cutoff.
    Groups consecutive rows and deletes each block in one operation."""
    to_delete = []
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=col).value
        if isinstance(val, datetime) and val < cutoff:
            to_delete.append(row_idx)

    if not to_delete:
        return 0

    ranges = []
    start = to_delete[0]
    end = to_delete[0]
    for r in to_delete[1:]:
        if r == end + 1:
            end = r
        else:
            ranges.append((start, end))
            start = end = r
    ranges.append((start, end))

    for start, end in reversed(ranges):
        ws.delete_rows(start, end - start + 1)

    return len(to_delete)

print("\n[4/7] Deleting IN/OUT rows before 2026...")
in_deleted = batch_delete_rows(ws_in, col=15, cutoff=CUTOFF)
print(f"  IN rows deleted: {in_deleted}, remaining: {ws_in.max_row}")
out_deleted = batch_delete_rows(ws_out, col=14, cutoff=CUTOFF)
print(f"  OUT rows deleted: {out_deleted}, remaining: {ws_out.max_row}")

# ============================================================
# Phase 5: Uncollapse/ungroup all row outlines in IN and OUT
#          (取消区域折叠)
# ============================================================
print("\n[5/7] Removing row grouping/outlining in IN and OUT...")

def ungroup_sheet(ws, name):
    """Remove all row grouping/outlining from a worksheet."""
    count = 0
    # Reset outline_level on every row that has it set
    for row_idx in range(1, ws.max_row + 1):
        rd = ws.row_dimensions.get(row_idx)
        if rd is not None and rd.outline_level is not None and rd.outline_level > 0:
            rd.outline_level = 0
            rd.hidden = False
            count += 1

    # Also clear sheet-level outline properties
    if ws.sheet_properties.outlinePr is not None:
        ws.sheet_properties.outlinePr.summaryBelow = True
        ws.sheet_properties.outlinePr.summaryRight = True

    print(f"  {name}: {count} rows ungrouped")

ungroup_sheet(ws_in, 'IN')
ungroup_sheet(ws_out, 'OUT')

# ============================================================
# Phase 6: Delete empty rows in IN and OUT
#          (清除没有数据的空白行)
# ============================================================
print("\n[6/7] Removing empty rows in IN and OUT...")

def delete_empty_rows(ws, name):
    """Delete rows that have no data at all (all cells None).
    Row 1 (header) is always kept."""
    to_delete = []
    for row_idx in range(2, ws.max_row + 1):
        # Check if the entire row is empty
        is_empty = True
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=row_idx, column=col_idx).value is not None:
                is_empty = False
                break
        if is_empty:
            to_delete.append(row_idx)

    if not to_delete:
        print(f"  {name}: no empty rows found")
        return 0

    # Group consecutive rows and batch delete
    ranges = []
    start = to_delete[0]
    end = to_delete[0]
    for r in to_delete[1:]:
        if r == end + 1:
            end = r
        else:
            ranges.append((start, end))
            start = end = r
    ranges.append((start, end))

    for start, end in reversed(ranges):
        ws.delete_rows(start, end - start + 1)

    print(f"  {name}: {len(to_delete)} empty rows deleted, {ws.max_row} remaining")
    return len(to_delete)

delete_empty_rows(ws_in, 'IN')
delete_empty_rows(ws_out, 'OUT')

# ============================================================
# Phase 7: Update Inventory columns Z-AF and Class
# ============================================================
print("\n[7/7] Updating Inventory scores and class...")

COL_UNIQUE = 26   # Z  — 唯一性
COL_Q      = 27   # AA — 质量 Quality
COL_S      = 28   # AB — 安全 Safety
COL_V      = 29   # AC — 价格 Value
COL_D      = 30   # AD — 交期 Delivery
COL_P_EQ   = 31   # AE — 所在设备重要性 Equipment importance
COL_SUM    = 32   # AF — 分值 Sum (formula)
COL_CLASS  = 12   # L  — 等级 class

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

rows_scored = 0
rows_red = 0

for info in inv_rows_info:
    row_idx = info['row']
    qty = info['current_qty']

    # Generate random 1-5 integers; reject until weighted sum > 4.5
    # Formula: Q*0.1 + S*0.1 + V*0.2 + D*0.3 + P*0.3
    for _ in range(10000):
        q_val = random.randint(1, 5)
        s_val = random.randint(1, 5)
        v_val = random.randint(1, 5)
        d_val = random.randint(1, 5)
        p_val = random.randint(1, 5)
        if q_val*0.1 + s_val*0.1 + v_val*0.2 + d_val*0.3 + p_val*0.3 > 4.5:
            break

    # Write integer values
    ws_inv.cell(row=row_idx, column=COL_UNIQUE).value = 0
    ws_inv.cell(row=row_idx, column=COL_Q).value = q_val
    ws_inv.cell(row=row_idx, column=COL_S).value = s_val
    ws_inv.cell(row=row_idx, column=COL_V).value = v_val
    ws_inv.cell(row=row_idx, column=COL_D).value = d_val
    ws_inv.cell(row=row_idx, column=COL_P_EQ).value = p_val

    # Excel formula for Sum
    formula = f"=AA{row_idx}*0.1+AB{row_idx}*0.1+AC{row_idx}*0.2+AD{row_idx}*0.3+AE{row_idx}*0.3"
    ws_inv.cell(row=row_idx, column=COL_SUM).value = formula

    # Set Class to "A"
    class_cell = ws_inv.cell(row=row_idx, column=COL_CLASS)
    class_cell.value = "A"

    # Red fill if stock quantity < 2
    if qty < 2:
        class_cell.fill = red_fill
        rows_red += 1

    rows_scored += 1

print(f"  Rows scored: {rows_scored}")
print(f"  Rows with red class (qty < 2): {rows_red}")

# ============================================================
# Save
# ============================================================
print(f"\nSaving to: {DST}")
wb.save(DST)
print("Done! over3.xlsx created successfully.")
