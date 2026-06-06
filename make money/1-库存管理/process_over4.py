"""
Process over3.xlsx → over4.xlsx

Operations:
1. Read Inventory data: Q(AA), S(AB), V(AC), D(AD), P(AE)
2. Compute Sum = Q*0.1 + S*0.1 + V*0.2 + D*0.3 + P*0.3
3. Assign grade(L) based on S value rules:
   - S > 4.5  → A
   - 2.5 < S < 4.5  → B
   - S < 2.5  → C
4. Write Excel formula for Sum in AF column
5. Derive Sum-based thresholds from S-graded data for future auto-classification
6. Save as over4.xlsx

Also provides a reusable classify_by_sum() function for new data entries.
"""
import openpyxl
from openpyxl.styles import PatternFill
from collections import defaultdict
import statistics

SRC = r'd:\【99】training code\make money\1-库存管理\over3.xlsx'
DST = r'd:\【99】training code\make money\1-库存管理\over4.xlsx'

# Column indices (1-based)
COL_MODEL  = 8    # H  — Model
COL_CLASS  = 12   # L  — 等级 class
COL_ORIG   = 15   # O  — Original quantity
COL_QTY    = 16   # P  — Quantity
COL_UNIQUE = 26   # Z  — 唯一性 unique
COL_Q      = 27   # AA — 质量 Quality
COL_S      = 28   # AB — 安全 Safety
COL_V      = 29   # AC — 价格 Value
COL_D      = 30   # AD — 交期 Delivery
COL_P_EQ   = 31   # AE — 所在设备重要性 Equipment importance
COL_SUM    = 32   # AF — 分值 Sum

# ============================================================
# Phase 1: Read current data
# ============================================================
print("[1/4] Reading over3.xlsx...")

# Read computed qty values in data_only mode (P column has formulas)
wb_data = openpyxl.load_workbook(SRC, data_only=True)
ws_inv_data = wb_data['Inventory']

# Build qty lookup: row_idx -> computed quantity value
qty_lookup = {}
empty_streak = 0
for row_idx in range(2, ws_inv_data.max_row + 1):
    model = ws_inv_data.cell(row=row_idx, column=COL_MODEL).value
    qty = ws_inv_data.cell(row=row_idx, column=COL_QTY).value
    if model is not None:
        try:
            qty_lookup[row_idx] = float(qty) if qty is not None else 0.0
        except (ValueError, TypeError):
            qty_lookup[row_idx] = 0.0
        empty_streak = 0
    else:
        empty_streak += 1
        if empty_streak >= 200:
            break
wb_data.close()

# Now read the editable workbook for Q/S/V/D/P values
wb = openpyxl.load_workbook(SRC)
ws_inv = wb['Inventory']

# Collect data rows (stop after 200 consecutive empty rows)
inv_rows = []
empty_streak = 0
for row_idx in range(2, ws_inv.max_row + 1):
    model = ws_inv.cell(row=row_idx, column=COL_MODEL).value
    if model is not None:
        inv_rows.append({
            'row': row_idx,
            'model': str(model).strip(),
            'q': ws_inv.cell(row=row_idx, column=COL_Q).value,
            's': ws_inv.cell(row=row_idx, column=COL_S).value,
            'v': ws_inv.cell(row=row_idx, column=COL_V).value,
            'd': ws_inv.cell(row=row_idx, column=COL_D).value,
            'p': ws_inv.cell(row=row_idx, column=COL_P_EQ).value,
            'qty': qty_lookup.get(row_idx, 0.0),
        })
        empty_streak = 0
    else:
        empty_streak += 1
        if empty_streak >= 200:
            break

print(f"  Inventory data rows: {len(inv_rows)}")

# ============================================================
# Phase 2: Compute Sum, assign grade by S rules, derive thresholds
# ============================================================
print("\n[2/4] Computing Sum and assigning grades by S-based rules...")

grade_sums = defaultdict(list)  # grade -> [sum_values]
grade_counts = defaultdict(int)
rows_updated = 0

def compute_sum(q, s, v, d, p):
    """Compute weighted sum from the 5 factors."""
    q = float(q) if q else 0
    s = float(s) if s else 0
    v = float(v) if v else 0
    d = float(d) if d else 0
    p = float(p) if p else 0
    return q * 0.1 + s * 0.1 + v * 0.2 + d * 0.3 + p * 0.3

def grade_by_s(s_val):
    """Determine grade based on S value.
    S > 4.5  → A
    2.5 < S < 4.5  → B
    S < 2.5  → C
    Boundary cases: S == 4.5 → B (not > 4.5), S == 2.5 → C (not > 2.5)
    """
    if s_val is None:
        return 'C'
    s = float(s_val)
    if s > 4.5:
        return 'A'
    elif s > 2.5:
        return 'B'
    else:
        return 'C'

red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

for info in inv_rows:
    row_idx = info['row']
    s_val = info['s']
    q_val = info['q']
    v_val = info['v']
    d_val = info['d']
    p_val = info['p']
    qty = info['qty']

    # Compute Sum
    sum_val = compute_sum(q_val, s_val, v_val, d_val, p_val)

    # Determine grade by S
    new_grade = grade_by_s(s_val)

    # Write Sum as Excel formula (not hardcoded value) for live recalculation
    formula = f"=AA{row_idx}*0.1+AB{row_idx}*0.1+AC{row_idx}*0.2+AD{row_idx}*0.3+AE{row_idx}*0.3"
    ws_inv.cell(row=row_idx, column=COL_SUM).value = formula

    # Write grade to L column
    class_cell = ws_inv.cell(row=row_idx, column=COL_CLASS)
    old_grade = class_cell.value
    class_cell.value = new_grade

    # Red fill if stock quantity < 2 (preserves previous behavior)
    if qty is not None and float(qty) < 2:
        class_cell.fill = red_fill

    if old_grade != new_grade:
        rows_updated += 1

    grade_sums[new_grade].append(sum_val)
    grade_counts[new_grade] += 1

print(f"  Grade distribution: A={grade_counts.get('A',0)}, B={grade_counts.get('B',0)}, C={grade_counts.get('C',0)}")
print(f"  Rows with grade changed: {rows_updated}")

# ============================================================
# Phase 3: Derive Sum-based thresholds for future auto-classification
# ============================================================
print("\n[3/4] Deriving Sum-based classification thresholds...")

# Use mean-based thresholds
mean_a = statistics.mean(grade_sums['A']) if grade_sums['A'] else 0
mean_b = statistics.mean(grade_sums['B']) if grade_sums['B'] else 0
mean_c = statistics.mean(grade_sums['C']) if grade_sums['C'] else 0

# Thresholds at midpoints between adjacent grade means
THRESHOLD_AB = round((mean_a + mean_b) / 2, 2)  # A/B boundary
THRESHOLD_BC = round((mean_b + mean_c) / 2, 2)  # B/C boundary

print(f"  Grade A mean Sum: {mean_a:.2f}")
print(f"  Grade B mean Sum: {mean_b:.2f}")
print(f"  Grade C mean Sum: {mean_c:.2f}")
print(f"  Sum threshold A/B: {THRESHOLD_AB}")
print(f"  Sum threshold B/C: {THRESHOLD_BC}")

# Validate accuracy
correct = 0
total = 0
for grade, sums in grade_sums.items():
    for s in sums:
        total += 1
        if s >= THRESHOLD_AB:
            pred = 'A'
        elif s >= THRESHOLD_BC:
            pred = 'B'
        else:
            pred = 'C'
        if pred == grade:
            correct += 1

print(f"  Sum-based classification accuracy: {correct}/{total} = {100*correct/total:.1f}%")

# Store thresholds in a hidden sheet for reference
if 'SumThresholds' in wb.sheetnames:
    del wb['SumThresholds']
ws_thresh = wb.create_sheet('SumThresholds')
ws_thresh['A1'] = 'Grade'
ws_thresh['B1'] = 'Sum Range'
ws_thresh['A2'] = 'A'
ws_thresh['B2'] = f'Sum >= {THRESHOLD_AB}'
ws_thresh['A3'] = 'B'
ws_thresh['B3'] = f'{THRESHOLD_BC} <= Sum < {THRESHOLD_AB}'
ws_thresh['A4'] = 'C'
ws_thresh['B4'] = f'Sum < {THRESHOLD_BC}'
ws_thresh['A6'] = 'Note'
ws_thresh['B6'] = 'Thresholds derived from S-based grading applied to current data. Sum = Q*0.1 + S*0.1 + V*0.2 + D*0.3 + P*0.3'

# ============================================================
# Phase 4: Save
# ============================================================
print(f"\n[4/4] Saving to: {DST}")
wb.save(DST)
print("Done! over4.xlsx created successfully.")

# ============================================================
# Print the classification function for reference
# ============================================================
print(f"""
# ============================================================
# Reusable classification function (also in process_over4.py)
# ============================================================
def classify_by_sum(q, s, v, d, p):
    '''
    Compute weighted Sum and return ABC grade.
    Sum = Q*0.1 + S*0.1 + V*0.2 + D*0.3 + P*0.3
    Grade thresholds derived from over3 data:
      Sum >= {THRESHOLD_AB} → A
      {THRESHOLD_BC} <= Sum < {THRESHOLD_AB} → B
      Sum < {THRESHOLD_BC} → C
    '''
    total = q*0.1 + s*0.1 + v*0.2 + d*0.3 + p*0.3
    if total >= {THRESHOLD_AB}:
        return 'A'
    elif total >= {THRESHOLD_BC}:
        return 'B'
    else:
        return 'C'
""")
